"""Dynamic reader for Kassirlar bo'yicha Excel report.

Supports variable column layouts — columns are discovered from the header rows
instead of relying on hard-coded positions.

KPI Excel column layout (Kassirlar.xlsx):
  Col 1:  №
  Col 2:  Табел рақами
  Col 3:  Лавозим
  Col 4:  Ф.И.Ш
  Col 5:  Изох (note/status)
  Col 6:  Локал код (branch MFO code)
  Col 7:  БХМ номи (branch name)
  Col 8:  Кассир иш хажми (юклама) — load %
  Col 9:  Иш куни — standard work days
  Col 10: Ишлаган куни — actual days worked
  Col 11: Operation-1 count (BEK)
  Col 12: Operation-1 minutes
  ...
  Col 39: Жами (Бек офис) count
  Col 40: Жами (Бек офис) minutes
  Col 41: Жами (Фронт офис) count
  Col 42: Жами (Фронт офис) minutes
"""
from __future__ import annotations
import json, re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import openpyxl
from openpyxl.utils import get_column_letter
from .db import _connect


def clean_branch_name(raw_branch: str) -> str:
    """Clean up MFO numbers and bank prefixes from branch names."""
    if not raw_branch:
        return ''
    s = str(raw_branch).strip()
    s = re.sub(r'^\d{5}\s*-\s*[0-9A-Za-z]{4,5}\s*', '', s)
    s = re.sub(r'^[0-9A-Za-z]{4,5}\s*-\s*', '', s)
    s = re.sub(r'^\d{5}\s*', '', s)
    s = re.sub(r'["\']O[^\s]*zsanoatqurilishbank["\']?\s*ATB\s*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'ATB\s*', '', s, flags=re.IGNORECASE)
    s = s.strip(' -"\'' + '\xa0')
    s = re.sub(r'bank\s+xizmatlar[i]?\s+markaz[i]?', 'BXM', s, flags=re.IGNORECASE)
    s = re.sub(r'bank\s+xizmatlar[i]?\s+ofis[i]?', 'BXO', s, flags=re.IGNORECASE)
    s = re.sub(r'центр\s+банковских\s+услуг', 'BXM', s, flags=re.IGNORECASE)
    s = re.sub(r'офис\s+банковских\s+услуг', 'BXO', s, flags=re.IGNORECASE)
    return s.strip()


def norm(v) -> str:
    """Normalise a cell value to a lowercase, ASCII-friendly string for matching."""
    if v is None:
        return ''
    s = str(v).lower().strip()
    # Cyrillic digraphs that appear in Uzbek sources
    cyr_map = {'қ': 'к', 'ў': 'у', 'ғ': 'г', 'ҳ': 'х', 'ё': 'е', 'ҷ': 'ч', 'ӣ': 'и'}
    for src, dst in cyr_map.items():
        s = s.replace(src, dst)
    s = re.sub(r"['`′ʼʻʼ]", '', s)
    s = re.sub(r'[^\w]+', ' ', s, flags=re.UNICODE)
    return s.strip()


def num(v) -> float:
    """Safely convert any cell value to a float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in ('-', '—', '- ', ' - ', 'None', 'null', 'nan'):
        return 0.0
    # Strip Excel non-breaking spaces and similar artefacts
    s = s.replace('\xa0', '').replace('\u202f', '')
    try:
        return float(s)
    except ValueError:
        cleaned = re.sub(r'[^0-9,.\-]', '', s).replace(',', '.')
        if not cleaned or cleaned in ('-', '.', '-.'):
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0


def sval(v) -> str:
    """Safely convert a cell value to a stripped string, removing Excel artefacts."""
    if v is None:
        return ''
    s = str(v).strip().replace('\xa0', ' ').replace('\u202f', ' ')
    # Remove leading/trailing apostrophes that openpyxl sometimes prepends
    s = s.strip("'")
    return s.strip()


# Keywords used to classify operation names into BEK (back-office) or FRONT groups.
BACK_KEYWORDS = (
    'амалга оширилган операциялар сони',
    'банкомат',
    'касса мудири',
    'кечки кассир',
    'назоратчи',
    'купюра санаш',
)
FRONT_KEYWORDS = (
    'ваш', 'валюта', 'коммунал', 'пластик', 'накд', 'кирим чиким хужжати',
)
SUMMARY_KEYWORDS = (
    'жами бек', 'jami bek', 'жами фронт', 'jami front',
    'жами амалиетлар', 'jami amaliyotlar',
    'бек фарк', 'фронт фарк', 'bek fark', 'front fark',
)

# Canonical display names for recognised operation types (used in metrics_json)
BACK_GROUPS = {
    'Амалга оширилган операциялар сони (кирим-чиқим)',
    'Банкоматга пул қўйиш',
    'Касса мудири',
    'Кечки кассир',
    'Назоратчи кассир ролини бажарганда',
    'Купюра санаш',
}
FRONT_GROUPS = {
    'Кирим, чиқим ҳужжатини текшириш, расмийлаштириш',
    'ВАЛЮТА 100$',
    'ВАЛЮТА 100,01–1000$',
    'ВАЛЮТА 1000,01–5000$',
    'ВАЛЮТА 10000$',
    'ВАЛЮТА 5000,01–10000$',
    'Коммунал тўловлар (кирим-чиқим)',
    'Пластик карта тарқатиш',
    'Пластикдан нақд пул ечиш',
}
EXCLUDED_GROUPS = {
    'БЕК фарқ', 'ФРОНТ фарқ', 'Жами (БЕК)', 'Жами (ФРОНТ)',
    'Жами амалиётлар', 'Жами',
    'БЕК фарк', 'ФРОНТ фарк', 'Жами БЕК', 'Жами ФРОНТ',
    'Жами (Бек офис)', 'Жами (Фронт офис)',
    'Бек офис', 'Фронт офис',
}


# ── HR status helpers ───────────────────────────────────────────────────────

def parse_hr_status(raw_note: str, position: str = ''):
    raw_note = str(raw_note or '').strip()
    position = str(position or '').strip()
    n_str = norm(f"{raw_note} {position}").replace(' ', '')

    if any(w in n_str for w in ('vacant', 'вакант', 'свобод', 'vakant', 'bosh', 'bush', 'бош', 'буш')):
        code, label = 'vacant', '⚪ Вакант (Свободная ставка)'
    elif any(w in n_str for w in ('dekret', 'декрет')):
        code, label = 'maternity', '🟣 В декрете'
    elif any(w in n_str for w in ('mexnat', 'мехнат', 'отпуск', 'tatil', 'татил')):
        code, label = 'vacation', '🔵 В отпуске (Меҳнат татили)'
    elif any(w in n_str for w in ('vakt', 'вакт', 'вактинча', 'замещ', 'вактинчалик', 'zamesh')):
        code, label = 'temporary', '🟡 Временный сотрудник'
    elif any(w in n_str for w in ('kasal', 'касал', 'больн')):
        code, label = 'sick', '🔴 Касал (Больничный)'
    else:
        code, label = 'active', '🟢 Работает'

    return {
        'status_code': code,
        'status_label': label,
        'has_replacement': 1 if code == 'active' else 0,
        'replacing_full_name': None,
        'replaced_by_full_name': None,
        'raw_note': raw_note,
    }


# ── DB schema initialisation ────────────────────────────────────────────────

def init_cashier_tables():
    with _connect() as c:
        c.execute("""CREATE TABLE IF NOT EXISTS cashier_imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            imported_at TEXT NOT NULL,
            rows_count INTEGER NOT NULL,
            report_label TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS cashier_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER NOT NULL REFERENCES cashier_imports(id) ON DELETE CASCADE,
            tab_number TEXT,
            full_name TEXT NOT NULL,
            position TEXT,
            days_worked REAL,
            operations_count REAL NOT NULL DEFAULT 0,
            operations_minutes REAL NOT NULL DEFAULT 0,
            bek_count REAL NOT NULL DEFAULT 0,
            bek_minutes REAL NOT NULL DEFAULT 0,
            front_count REAL NOT NULL DEFAULT 0,
            front_minutes REAL NOT NULL DEFAULT 0,
            branch_name TEXT,
            raw_note TEXT,
            hr_status_code TEXT,
            hr_status_label TEXT,
            replacing_full_name TEXT,
            replaced_by_full_name TEXT,
            has_replacement INTEGER DEFAULT 1,
            metrics_json TEXT NOT NULL DEFAULT '{}'
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS cashier_status_imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            imported_at TEXT NOT NULL,
            rows_count INTEGER NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS cashier_statuses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER NOT NULL REFERENCES cashier_status_imports(id) ON DELETE CASCADE,
            branch_name TEXT,
            position TEXT,
            full_name TEXT NOT NULL,
            status_code TEXT NOT NULL,
            status_label TEXT NOT NULL,
            raw_note TEXT,
            replacing_full_name TEXT,
            replaced_by_full_name TEXT,
            has_replacement INTEGER NOT NULL DEFAULT 0
        )""")

        # Migrations — add columns that may be absent in older databases
        for table, col, definition in [
            ('cashier_reports', 'branch_name', 'TEXT'),
            ('cashier_reports', 'raw_note', 'TEXT'),
            ('cashier_reports', 'hr_status_code', 'TEXT'),
            ('cashier_reports', 'hr_status_label', 'TEXT'),
            ('cashier_reports', 'replacing_full_name', 'TEXT'),
            ('cashier_reports', 'replaced_by_full_name', 'TEXT'),
            ('cashier_reports', 'has_replacement', 'INTEGER DEFAULT 1'),
        ]:
            cols = [col_info[1] for col_info in c.execute(f"PRAGMA table_info({table})").fetchall()]
            if col not in cols:
                c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {definition}")


# ── KPI Excel parser ────────────────────────────────────────────────────────

def _classify_op_group(op_name: str) -> str:
    """Classify an operation name into 'back', 'front', or 'summary'."""
    n = norm(op_name)
    if any(kw in n for kw in SUMMARY_KEYWORDS):
        return 'summary'
    if any(kw in n for kw in BACK_KEYWORDS):
        return 'back'
    return 'front'


def _find_header_rows(all_rows: list) -> Tuple[int, int, int]:
    """
    Scan the first 30 rows to find the main header row index (0-based),
    optional sub-header row with 'Сони'/'минутда', and the data start row index.
    Returns (header_idx, sub_header_idx_or_-1, data_start_idx).
    """
    header_idx = None
    for r_idx in range(min(30, len(all_rows))):
        row = all_rows[r_idx]
        row_text = norm(' '.join(str(c or '') for c in row))
        # Look for characteristic FIO/tab column headers
        if any(k in row_text for k in ('ф и ш', 'фиш', 'ф и о', 'фио', 'fish', 'fio', 'табел')):
            header_idx = r_idx
            break

    if header_idx is None:
        header_idx = 1  # fallback: second row

    # Check if the next row is a sub-header (Сони / минутда labels)
    sub_header_idx = -1
    if header_idx + 1 < len(all_rows):
        next_row = all_rows[header_idx + 1]
        next_text = norm(' '.join(str(c or '') for c in next_row))
        if any(w in next_text for w in ('сони', 'минут', 'count', 'minut')):
            sub_header_idx = header_idx + 1

    data_start_idx = sub_header_idx + 1 if sub_header_idx >= 0 else header_idx + 1
    # Skip a possible "totals" row (row 4 in Kassirlar.xlsx which has grand totals)
    if data_start_idx < len(all_rows):
        totals_row = all_rows[data_start_idx]
        non_empty = [v for v in totals_row if v not in (None, '')]
        # If the first non-empty cell is a large integer (≥ 1000) it's a totals row
        first_val = non_empty[0] if non_empty else None
        if isinstance(first_val, (int, float)) and first_val >= 1000:
            data_start_idx += 1

    return header_idx, sub_header_idx, data_start_idx


def _build_op_column_map(all_rows: list, header_idx: int, sub_header_idx: int) -> Dict:
    """
    Build a mapping of operation columns from the multi-row header.

    Returns a dict:
      {
        'fixed': {
            'num': col_idx,    # № column (0-based)
            'tab': col_idx,
            'pos': col_idx,
            'fio': col_idx,
            'note': col_idx,
            'bcode': col_idx,  # branch MFO code
            'bname': col_idx,  # branch name
            'load': col_idx,   # load %
            'std_days': col_idx,   # standard work days
            'days': col_idx,   # actual days worked
            'bek_total_cnt': col_idx,
            'bek_total_min': col_idx,
            'front_total_cnt': col_idx,
            'front_total_min': col_idx,
        },
        'ops': [
            {'name': str, 'group': 'back'|'front'|'other', 'cnt_col': int, 'min_col': int},
            ...
        ]
      }
    """
    header_row = all_rows[header_idx]
    sub_row = all_rows[sub_header_idx] if sub_header_idx >= 0 else [None] * len(header_row)

    # Expand header: when a cell is None in sub_header, carry over the parent header label
    # to get the full col-level name.
    n_cols = max(len(header_row), len(sub_row))
    header_row = list(header_row) + [None] * (n_cols - len(header_row))
    sub_row = list(sub_row) + [None] * (n_cols - len(sub_row))

    # Carry-forward logic for merged header cells across columns
    cur_group = ''
    col_label = [''] * n_cols  # final label for each column
    col_is_soni = [False] * n_cols
    col_is_minut = [False] * n_cols

    for idx in range(n_cols):
        hv = sval(header_row[idx])
        sv = sval(sub_row[idx])
        nh = norm(hv)
        ns = norm(sv)

        if hv:
            cur_group = hv
        col_label[idx] = cur_group

        if 'сони' in ns or 'count' in ns or ('soni' in ns and 'ish' not in ns):
            col_is_soni[idx] = True
        if 'минут' in ns or 'minut' in ns:
            col_is_minut[idx] = True

    # Now parse fixed info columns (non-operation columns before operations start)
    fixed = {}
    for idx, hv in enumerate(header_row):
        hv_s = sval(hv)
        nh = norm(hv_s).replace(' ', '')
        if not nh:
            continue

        if '№' in hv_s or nh in ('', 'nomer', '№') and idx == 0:
            fixed.setdefault('num', idx)
        elif any(k in nh for k in ('табел', 'tabel')):
            fixed.setdefault('tab', idx)
        elif any(k in nh for k in ('лавозим', 'lavozim', 'должность', 'position')):
            fixed.setdefault('pos', idx)
        elif any(k in nh for k in ('ф и ш', 'фиш', 'ф и о', 'фио', 'fish', 'fio', 'f i sh')):
            fixed.setdefault('fio', idx)
        elif any(k in nh for k in ('изох', 'izox', 'примечание', 'note', 'статус')):
            fixed.setdefault('note', idx)
        elif any(k in nh for k in ('локалкод', 'мфо', 'mfo', 'кодфилиал', 'localcod')):
            fixed.setdefault('bcode', idx)
        elif any(k in nh for k in ('бхмноми', 'bhm', 'филиалноми', 'наименованиефилиал', 'bhmnomi')):
            fixed.setdefault('bname', idx)
        elif any(k in nh for k in ('юклама', 'yuklama', 'хажми', 'hazmi')):
            fixed.setdefault('load', idx)
        elif any(k in nh for k in ('ишкуни', 'ishkuni')) and 'ишлаган' not in nh and 'ishlagan' not in nh:
            fixed.setdefault('std_days', idx)
        elif any(k in nh for k in ('ишлагакуни', 'ишлаганкуни', 'ishlagankuni', 'ishlagan')):
            fixed.setdefault('days', idx)

    # Fallback if dynamic detection fails (for Kassirlar.xlsx standard layout)
    fixed.setdefault('num', 0)
    fixed.setdefault('tab', 1)
    fixed.setdefault('pos', 2)
    fixed.setdefault('fio', 3)
    fixed.setdefault('note', 4)
    fixed.setdefault('bcode', 5)
    fixed.setdefault('bname', 6)
    fixed.setdefault('load', 7)
    fixed.setdefault('std_days', 8)
    fixed.setdefault('days', 9)

    # Find operation pair columns (Сони + минутда pairs)
    # and special summary columns (Жами Бек офис, Жами Фронт офис)
    ops = []
    bek_total_cnt = None
    bek_total_min = None
    front_total_cnt = None
    front_total_min = None

    # Operation pair scan: look for consecutive Сони/минутда pairs starting from col 10
    first_op_col = max(fixed['days'] + 1, 10)

    i = first_op_col
    while i < n_cols:
        label = col_label[i]
        n_label = norm(label)

        # Detect summary columns
        if any(k in n_label for k in ('жами бек офис', 'jami bek ofis', 'жами бек', 'jami bek')):
            if col_is_soni[i]:
                bek_total_cnt = i
            elif col_is_minut[i]:
                bek_total_min = i
            # Check adjacent
            if i + 1 < n_cols and col_is_minut[i + 1] and norm(col_label[i + 1]) == n_label:
                if bek_total_cnt is None:
                    bek_total_cnt = i
                bek_total_min = i + 1
                i += 2
                continue
        elif any(k in n_label for k in ('жами фронт офис', 'jami front ofis', 'жами фронт', 'jami front')):
            if col_is_soni[i]:
                front_total_cnt = i
            elif col_is_minut[i]:
                front_total_min = i
            if i + 1 < n_cols and col_is_minut[i + 1] and norm(col_label[i + 1]) == n_label:
                if front_total_cnt is None:
                    front_total_cnt = i
                front_total_min = i + 1
                i += 2
                continue

        # Regular operation: must have a non-empty label and be a Сони col
        if label and col_is_soni[i] and not any(k in n_label for k in ('жами', 'jami')):
            cnt_col = i
            min_col = i + 1 if (i + 1 < n_cols and col_is_minut[i + 1]) else None
            group = _classify_op_group(label)
            ops.append({
                'name': label,
                'group': group,
                'cnt_col': cnt_col,
                'min_col': min_col,
            })
            i += 2 if min_col is not None else 1
            continue

        i += 1

    # If summary cols not yet found, try last two pairs
    if bek_total_cnt is None:
        # Heuristic: last-4 columns should be: bek_cnt, bek_min, front_cnt, front_min
        candidate = n_cols - 4
        if candidate > first_op_col:
            bek_total_cnt = candidate
            bek_total_min = candidate + 1
            front_total_cnt = candidate + 2
            front_total_min = candidate + 3

    fixed['bek_total_cnt'] = bek_total_cnt
    fixed['bek_total_min'] = bek_total_min
    fixed['front_total_cnt'] = front_total_cnt
    fixed['front_total_min'] = front_total_min

    return {'fixed': fixed, 'ops': ops}


def parse_cashiers_xlsx(path) -> dict:
    """Parse KPI Excel report. Returns dict with 'records' list and metadata."""
    path = Path(path)
    path_str = str(path)

    if path_str.lower().endswith('.csv'):
        import csv
        with open(path, 'r', encoding='utf-8-sig', errors='ignore') as f:
            all_rows = [tuple(row) for row in csv.reader(f)]
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

    if not all_rows:
        raise ValueError('Excel-файл пуст.')

    header_idx, sub_header_idx, data_start_idx = _find_header_rows(all_rows)
    col_map = _build_op_column_map(all_rows, header_idx, sub_header_idx)
    fx = col_map['fixed']
    ops_def = col_map['ops']

    records = []
    errors = []
    SKIP_NAMES = frozenset({
        'жами', 'итого', 'total', 'всего', 'сони', 'минут',
        'фиш', 'фио', 'ф и ш', 'ф и о', 'жами сони',
    })
    SKIP_TABS = frozenset({'жами', 'итого', 'total', 'всего', 'номер', '№'})

    for rn in range(data_start_idx, len(all_rows)):
        row = all_rows[rn]
        if not row or not any(v not in (None, '') for v in row):
            continue

        def get(col_idx):
            """Get cell value by 0-based column index."""
            if col_idx is None or col_idx >= len(row):
                return None
            return row[col_idx]

        name = sval(get(fx['fio']))
        tab_num = sval(get(fx['tab'])) or sval(get(fx['num']))
        pos = sval(get(fx['pos']))
        note = sval(get(fx['note'])) if fx.get('note') is not None else ''
        bcode = sval(get(fx['bcode'])) if fx.get('bcode') is not None else ''
        bname = sval(get(fx['bname'])) if fx.get('bname') is not None else ''

        # Skip summary/header rows embedded in data
        n_name = norm(name).replace(' ', '')
        n_tab = norm(tab_num).replace(' ', '')
        if n_name in SKIP_NAMES or n_tab in SKIP_TABS:
            continue
        if not name or name in ('-', '—', 'None', 'null', 'nan'):
            continue
        # Skip if name is purely numeric (row number column)
        if re.match(r'^\d+$', name):
            continue

        # ── Fixed KPI fields ────────────────────────────────────────────
        load_raw = num(get(fx['load']))
        # Normalise: if stored as ratio (0-1), convert to percent
        if 0 < load_raw <= 1.0:
            load_pct = round(load_raw * 100, 1)
        else:
            load_pct = round(load_raw, 1)

        days_worked = num(get(fx['days'])) or num(get(fx.get('std_days')))

        # ── Branch ─────────────────────────────────────────────────────
        if bcode and bname:
            branch_str = f"{bcode} - {bname}"
        elif bname:
            branch_str = bname
        elif bcode:
            branch_str = bcode
        else:
            branch_str = ''

        # ── HR status from note / position ─────────────────────────────
        hr = parse_hr_status(note, pos)

        # ── Operation metrics ──────────────────────────────────────────
        metrics = {}
        bek_ops_cnt = 0.0
        bek_ops_min = 0.0
        front_ops_cnt = 0.0
        front_ops_min = 0.0

        for op in ops_def:
            cnt = num(get(op['cnt_col']))
            mins = num(get(op['min_col'])) if op['min_col'] is not None else 0.0
            op_name = op['name']
            metrics[op_name] = {'count': cnt, 'minutes': mins}
            if op['group'] == 'back':
                bek_ops_cnt += cnt
                bek_ops_min += mins
            elif op['group'] == 'front':
                front_ops_cnt += cnt
                front_ops_min += mins

        # ── Summary totals from dedicated columns (preferred) ───────────
        bek_cnt_raw = num(get(fx.get('bek_total_cnt')))
        bek_min_raw = num(get(fx.get('bek_total_min')))
        front_cnt_raw = num(get(fx.get('front_total_cnt')))
        front_min_raw = num(get(fx.get('front_total_min')))

        # Use dedicated summary cols if available, else fall back to summed ops
        bek_count = bek_cnt_raw if bek_cnt_raw > 0 else bek_ops_cnt
        bek_minutes = bek_min_raw if bek_min_raw > 0 else bek_ops_min
        front_count = front_cnt_raw if front_cnt_raw > 0 else front_ops_cnt
        front_minutes = front_min_raw if front_min_raw > 0 else front_ops_min

        ops_count = bek_count + front_count
        ops_minutes = bek_minutes + front_minutes

        # Recompute load_percent if not available from header
        if load_pct == 0 and days_worked > 0 and ops_minutes > 0:
            load_pct = round((ops_minutes / (days_worked * 480)) * 100, 1)

        rec = {
            'full_name': name,
            'tab_number': tab_num,
            'position': pos,
            'days_worked': days_worked,
            'branch_name': branch_str,
            'raw_note': note,
            'hr_status_code': hr['status_code'],
            'hr_status_label': hr['status_label'],
            'has_replacement': hr['has_replacement'],
            'replacing_full_name': None,
            'replaced_by_full_name': None,
            'operations_count': ops_count,
            'operations_minutes': ops_minutes,
            'bek_count': bek_count,
            'bek_minutes': bek_minutes,
            'front_count': front_count,
            'front_minutes': front_minutes,
            'load_percent': load_pct,
            'load_difference': 0.0,
            'employee_number': sval(get(fx['num'])),
            'metrics': metrics,
        }
        records.append(rec)

    if not records:
        raise ValueError(
            f'После строки шапки {header_idx + 1} не найдено валидных строк кассиров.'
        )

    # ── Rule 3: Adjacent correlation (вакт. directly above декрет/мехнат.тат.) ─
    for i in range(len(records) - 1):
        r_curr = records[i]
        r_next = records[i + 1]
        if (r_curr['hr_status_code'] == 'temporary'
                and r_next['hr_status_code'] in ('maternity', 'vacation')
                and not r_curr.get('replacing_full_name')
                and not r_next.get('replaced_by_full_name')):
            r_curr['replacing_full_name'] = r_next['full_name']
            r_next['replaced_by_full_name'] = r_curr['full_name']
            r_next['has_replacement'] = 1
            r_curr['hr_status_label'] = f"🟡 Временный (замещает {r_next['full_name']})"
            if r_next['hr_status_code'] == 'maternity':
                r_next['hr_status_label'] = f"🟣 В декрете (замещает {r_curr['full_name']})"
            elif r_next['hr_status_code'] == 'vacation':
                r_next['hr_status_label'] = f"🔵 В отпуске (замещает {r_curr['full_name']})"

    # ── Rule 4: Filial-group correlation ───────────────────────────────────
    branch_groups: dict[str, list] = {}
    for r in records:
        b_key = r.get('branch_name') or 'Default'
        branch_groups.setdefault(b_key, []).append(r)

    for b_key, b_recs in branch_groups.items():
        unmatched_temps = [r for r in b_recs if r['hr_status_code'] == 'temporary' and not r.get('replacing_full_name')]
        unmatched_absents = [r for r in b_recs if r['hr_status_code'] in ('maternity', 'vacation') and not r.get('replaced_by_full_name')]
        for temp_r, abs_r in zip(unmatched_temps, unmatched_absents):
            temp_r['replacing_full_name'] = abs_r['full_name']
            abs_r['replaced_by_full_name'] = temp_r['full_name']
            abs_r['has_replacement'] = 1
            temp_r['hr_status_label'] = f"🟡 Временный (замещает {abs_r['full_name']})"
            if abs_r['hr_status_code'] == 'maternity':
                abs_r['hr_status_label'] = f"🟣 В декрете (замещает {temp_r['full_name']})"
            elif abs_r['hr_status_code'] == 'vacation':
                abs_r['hr_status_label'] = f"🔵 В отпуске (замещает {temp_r['full_name']})"

        # Remaining absents without replacement
        for abs_r in b_recs:
            if abs_r['hr_status_code'] in ('maternity', 'vacation') and not abs_r.get('replaced_by_full_name'):
                abs_r['has_replacement'] = 0
                if abs_r['hr_status_code'] == 'maternity':
                    abs_r['hr_status_label'] = '🟣 В декрете (без замены!)'
                elif abs_r['hr_status_code'] == 'vacation':
                    abs_r['hr_status_label'] = '🔵 В отпуске (без замены!)'

    return {
        'records': records,
        'errors': errors,
        'header_rows': [header_idx + 1, sub_header_idx + 1 if sub_header_idx >= 0 else header_idx + 1],
        'columns': [op['name'] for op in ops_def],
    }


def save_cashier_import(filename, parsed):
    init_cashier_tables()
    with _connect() as c:
        iid = c.execute(
            'INSERT INTO cashier_imports(filename,imported_at,rows_count) VALUES(?,?,?)',
            (filename, datetime.now(timezone.utc).isoformat(), len(parsed['records']))
        ).lastrowid
        for r in parsed['records']:
            c.execute('''INSERT INTO cashier_reports(
                import_id, tab_number, full_name, position, days_worked,
                operations_count, operations_minutes, bek_count, bek_minutes,
                front_count, front_minutes, branch_name, raw_note,
                hr_status_code, hr_status_label, replacing_full_name, replaced_by_full_name,
                has_replacement, metrics_json
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (
                iid,
                r.get('tab_number'),
                r['full_name'],
                r.get('position'),
                r.get('days_worked', 0),
                r.get('operations_count', 0),
                r.get('operations_minutes', 0),
                r.get('bek_count', 0),
                r.get('bek_minutes', 0),
                r.get('front_count', 0),
                r.get('front_minutes', 0),
                r.get('branch_name', ''),
                r.get('raw_note', ''),
                r.get('hr_status_code', 'active'),
                r.get('hr_status_label', '🟢 Работает'),
                r.get('replacing_full_name'),
                r.get('replaced_by_full_name'),
                r.get('has_replacement', 1),
                json.dumps({
                    'employee_number': r.get('employee_number'),
                    'total_marker': r.get('tab_number'),
                    'load_percent': r.get('load_percent', 0),
                    'load_difference': r.get('load_difference', 0),
                    'operations': r.get('metrics', {}),
                }, ensure_ascii=False),
            ))
    return {'import_id': iid, 'imported': len(parsed['records']), 'header_rows': parsed['header_rows']}


# ── Shtat (HR status) Excel parser ─────────────────────────────────────────

def parse_cashier_status_xlsx(path) -> dict:
    """Parse the HR staff registry Excel (Штат.xlsx)."""
    path = Path(path)
    path_str = str(path)

    if path_str.lower().endswith('.csv'):
        import csv
        with open(path, 'r', encoding='utf-8-sig', errors='ignore') as f:
            all_rows = [tuple(row) for row in csv.reader(f)]
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

    if not all_rows:
        raise ValueError('Excel-файл пуст.')

    # ── Column detection ────────────────────────────────────────────────────
    fio_col = None
    pos_col = None
    mfo_col = None
    note_col = None
    header_row_idx = None

    for r_idx in range(min(30, len(all_rows))):
        row = all_rows[r_idx]
        if not row:
            continue
        r_vals = [sval(x) for x in row]
        n_vals = [norm(x) for x in r_vals]

        f_c = p_c = m_c = n_c = None
        for idx, nv in enumerate(n_vals):
            nv_nospace = nv.replace(' ', '')
            if any(w in nv_nospace for w in ('фиш', 'фио', 'fio', 'fish', 'сотрудник', 'работник', 'xodim', 'ф и ш', 'ф и о')):
                f_c = idx
            elif any(w in nv_nospace for w in ('таркибий', 'лавозим', 'должность', 'lavozim', 'position')) and \
                    not any(k in nv_nospace for k in ('маош', 'оклад', 'разряд', 'коэффициент')):
                p_c = idx
            elif any(w in nv_nospace for w in ('mfo', 'мфо', 'кодфилиала')):
                m_c = idx
            elif any(w in nv_nospace for w in ('изох', 'примечание', 'статус', 'причина', 'note')):
                n_c = idx

        if f_c is not None or p_c is not None:
            fio_col, pos_col, mfo_col, note_col = f_c, p_c, m_c, n_c
            header_row_idx = r_idx
            break

    # Fallback for Штат.xlsx standard layout:
    # Col 0=№, 1=MFO, 2=Lavozim(position), 3=Штат бирликлари, 4=Разряд,
    # 5=Коэффициент, 6=Оклад, 7=Ф.И.Ш, 8=Ставка, 9=Буш урни, 10=Статус
    if fio_col is None:
        fio_col = 7
    if pos_col is None:
        pos_col = 2
    if mfo_col is None:
        mfo_col = 1
    if note_col is None:
        note_col = 10  # Status column in Штат.xlsx

    # We store (row_index_in_all_rows, row_values) to enable adjacency pairing
    raw_parsed = []  # list of (rn, dict) to preserve order
    current_branch = 'Не указан'
    current_pos = ''
    current_mfo = ''

    start_idx = (header_row_idx + 1) if header_row_idx is not None else 0

    SKIP_FIO = frozenset({
        'ф и ш', 'ф и о', 'фио', 'фиш', 'jami', 'итого', 'сони', 'минут',
        'номер', '№', 'ставка', 'всего', 'жами',
    })

    for rn in range(start_idx, len(all_rows)):
        row = all_rows[rn]
        if not row or not any(v not in (None, '') for v in row):
            continue

        r_str = [sval(cell) for cell in row]

        def gcol(idx, r=r_str):
            return r[idx] if idx is not None and idx < len(r) else ''

        first_val = r_str[0] if r_str else ''
        n_first = norm(first_val)

        # ── Track region/branch headers ────────────────────────────────
        if any(w in n_first for w in ('регион', 'region')):
            continue
        if first_val.startswith('Филиал:') or any(w in n_first for w in ('филиал', 'бхм', 'бхо', 'центр', 'офис')):
            current_branch = first_val.replace('Филиал:', '').strip()
            current_mfo = gcol(mfo_col)
            continue
        if first_val.startswith('**'):
            current_pos = first_val.replace('**', '').strip()
            continue
        if any(w in n_first for w in ('жами:', 'жами', 'итого', 'total')) and len(first_val) < 20:
            continue

        # ── Extract values ─────────────────────────────────────────────
        fio_val = gcol(fio_col)
        pos_val = gcol(pos_col)
        mfo_val = gcol(mfo_col)
        note_val = gcol(note_col)

        # is_subrow: position cell is empty — this row belongs to the previous main row's position slot
        is_subrow = not pos_val.strip() or pos_val.strip() in ('-', '—')

        if not is_subrow:
            current_pos = pos_val

        # Carry-forward position and MFO
        if is_subrow:
            pos_val = current_pos
        if not mfo_val or mfo_val in ('-', '—'):
            mfo_val = current_mfo
        else:
            current_mfo = mfo_val

        # Fallback FIO detection
        if not fio_val:
            for test_idx in (7, 6, 5, 3, 2):
                val = r_str[test_idx] if len(r_str) > test_idx else ''
                n_v = norm(val)
                if val and len(val) >= 3 and n_v not in SKIP_FIO:
                    if 'вакант' in n_v or 'vakant' in n_v or len(val.split()) >= 2:
                        fio_val = val
                        break

        n_fio = norm(fio_val)
        if not fio_val or len(fio_val) < 3 or n_fio in SKIP_FIO:
            continue

        is_vacant = any(w in n_fio for w in ('вакант', 'vakant', 'bosh', 'буш', 'bush'))
        full_name = 'ВАКАНТ' if is_vacant else fio_val

        mfo_clean = mfo_val.replace('\xa0', '').strip()
        branch_label = current_branch
        if mfo_clean and mfo_clean not in current_branch:
            branch_label = f"{mfo_clean} - {current_branch}"

        clean_note = note_val.replace('\xa0', '').strip()
        if clean_note.isdigit():
            clean_note = ''

        raw_parsed.append((rn, is_subrow, {
            'branch_name': branch_label,
            'position': pos_val if pos_val else 'Кассир',
            'full_name': full_name,
            'raw_note': clean_note,
            'status_code': 'vacant' if is_vacant else 'active',
            'status_label': '⚪ Вакант' if is_vacant else '🟢 Ишлаяпти',
            'replacing_full_name': None,
            'replaced_by_full_name': None,
            'has_replacement': 0,
        }))

    records = [rec for _, _, rec in raw_parsed]

    if not records:
        raise ValueError('Не найдено ни одной валидной строки с ФИО кассиров или вакансиями в реестре штата.')

    # ── Pass 1: Determine status_code from raw_note ─────────────────────────
    for r in records:
        if r['status_code'] == 'vacant':
            continue
        n_flat = norm(f"{r['position']} {r['raw_note']}").replace(' ', '')
        if 'dekret' in n_flat or 'декрет' in n_flat:
            r['status_code'] = 'maternity'
        elif any(w in n_flat for w in ('mexnattat', 'мехнатtat', 'мехнаттатил', 'mexnattatil', 'татил', 'tatil')):
            r['status_code'] = 'vacation'
        elif any(w in n_flat for w in ('mexnat', 'мехнат', 'отпуск')):
            r['status_code'] = 'vacation'
        elif any(w in n_flat for w in ('vakt', 'вакт', 'zamesh', 'замещ')):
            r['status_code'] = 'temporary'
        elif any(w in n_flat for w in ('kasal', 'касал', 'больн')):
            r['status_code'] = 'sick'

    # ── Pass 2: Sub-row adjacency pairing (highest priority) ────────────────
    # In Штат.xlsx, when a temp (вакт.) employee is in row N,
    # the absent (декрет/мехнат) employee they replace is in the very next
    # row N+1 as a 'sub-row' (position cell is empty).
    # This is the canonical, most accurate pairing signal.
    for i in range(len(raw_parsed) - 1):
        _, is_sub_curr, r_curr = raw_parsed[i]
        _, is_sub_next, r_next = raw_parsed[i + 1]

        # Pattern 1: temp in row i, absent in sub-row i+1
        if (r_curr['status_code'] == 'temporary'
                and is_sub_next
                and r_next['status_code'] in ('maternity', 'vacation')
                and not r_curr.get('replacing_full_name')
                and not r_next.get('replaced_by_full_name')):
            r_curr['replacing_full_name'] = r_next['full_name']
            r_next['replaced_by_full_name'] = r_curr['full_name']
            r_next['has_replacement'] = 1

        # Pattern 2: absent in sub-row i, temp in adjacent row i+1
        elif (is_sub_curr
                and r_curr['status_code'] in ('maternity', 'vacation')
                and r_next['status_code'] == 'temporary'
                and not r_curr.get('replaced_by_full_name')
                and not r_next.get('replacing_full_name')):
            r_next['replacing_full_name'] = r_curr['full_name']
            r_curr['replaced_by_full_name'] = r_next['full_name']
            r_curr['has_replacement'] = 1

    # ── Pass 3: Within same branch — match remaining unmatched temps to absents ─
    branch_recs: dict[str, list] = {}
    for r in records:
        bk = r.get('branch_name') or 'Default'
        branch_recs.setdefault(bk, []).append(r)

    for bk, b_list in branch_recs.items():
        # Only pair within same position type for accuracy
        pos_groups: dict[str, list] = {}
        for r in b_list:
            pg = norm(r.get('position', '')).replace(' ', '')
            pos_groups.setdefault(pg, []).append(r)

        for pg, pg_list in pos_groups.items():
            unmatched_temps = [r for r in pg_list if r['status_code'] == 'temporary' and not r.get('replacing_full_name')]
            unmatched_absents = [r for r in pg_list if r['status_code'] in ('maternity', 'vacation') and not r.get('replaced_by_full_name')]
            for temp_r, abs_r in zip(unmatched_temps, unmatched_absents):
                temp_r['replacing_full_name'] = abs_r['full_name']
                abs_r['replaced_by_full_name'] = temp_r['full_name']
                abs_r['has_replacement'] = 1

        # Mark remaining absents without any replacement
        for r in b_list:
            if r['status_code'] in ('maternity', 'vacation') and not r.get('replaced_by_full_name'):
                r['has_replacement'] = 0

    # ── Pass 4: Build readable labels ──────────────────────────────────────
    for r in records:
        sc = r['status_code']
        if sc == 'vacant':
            r['status_label'] = '⚪ Вакант (Свободная ставка)'
        elif sc == 'temporary':
            r['status_label'] = (
                f"🟡 Вақтинча (ўринбосар: {r['replacing_full_name']})"
                if r.get('replacing_full_name') else '🟡 Вақтинча ходим'
            )
        elif sc == 'maternity':
            r['status_label'] = (
                f"🟣 Декретда (ўрнида: {r['replaced_by_full_name']})"
                if r.get('replaced_by_full_name') else '🟣 Декретда (без замены!)'
            )
        elif sc == 'vacation':
            r['status_label'] = (
                f"🔵 Меҳнат татилда (ўрнида: {r['replaced_by_full_name']})"
                if r.get('replaced_by_full_name') else '🔵 Меҳнат татилда (без замены!)'
            )
        elif sc == 'sick':
            r['status_label'] = '🔴 Касал (Больничный)'
        else:
            r['status_code'] = 'active'
            r['status_label'] = '🟢 Ишлаяпти'
            r['has_replacement'] = 1

    return {'records': records, 'total_rows': len(all_rows)}


def save_cashier_status_import(filename, parsed):
    init_cashier_tables()
    with _connect() as c:
        iid = c.execute(
            'INSERT INTO cashier_status_imports(filename,imported_at,rows_count) VALUES(?,?,?)',
            (filename, datetime.now(timezone.utc).isoformat(), len(parsed['records']))
        ).lastrowid
        for r in parsed['records']:
            c.execute(
                '''INSERT INTO cashier_statuses(
                    import_id, branch_name, position, full_name, status_code, status_label,
                    raw_note, replacing_full_name, replaced_by_full_name, has_replacement
                ) VALUES(?,?,?,?,?,?,?,?,?,?)''',
                (
                    iid, r.get('branch_name'), r.get('position'), r['full_name'],
                    r['status_code'], r['status_label'], r.get('raw_note'),
                    r.get('replacing_full_name'), r.get('replaced_by_full_name'),
                    r['has_replacement'],
                )
            )
    return {'import_id': iid, 'imported': len(parsed['records'])}


# ── Analytics helpers ───────────────────────────────────────────────────────

def cashier_role(row) -> str:
    b = float(row.get('bek_count', 0) or 0)
    f = float(row.get('front_count', 0) or 0)
    if b == 0 and f == 0:
        return 'back'
    return 'front' if f > b else 'back'


def compute_efficiency_score(x: dict) -> dict:
    lp = float(x.get('load_percent', 0) or 0)
    diff = abs(float(x.get('load_difference', 0) or 0))

    # Workload score (max 35)
    s_load = round(35.0 * min(lp, 100.0) / 100.0, 1)
    if diff > 0:
        s_load = max(0.0, round(s_load - min(15.0, diff * 5.0), 1))

    ops = float(x.get('operations_count', 0) or 0)
    s_ops = round(min(30.0, 30.0 * (ops / 150.0)), 1) if ops else 0.0

    sec = float(x.get('avg_seconds_per_operation', 0) or 0)
    if 0 < sec <= 180:
        s_speed = 20.0
    elif sec > 180:
        s_speed = round(max(5.0, 20.0 - (sec - 180) * 0.05), 1)
    else:
        s_speed = 0.0

    days = float(x.get('days_worked', 0) or 0)
    s_days = round(min(15.0, 15.0 * (days / 22.0)), 1)

    total_score = round(s_load + s_ops + s_speed + s_days, 1)
    if total_score >= 85:
        grade = 'Top Performer'
    elif total_score >= 70:
        grade = 'Optimal'
    elif total_score >= 50:
        grade = 'Average'
    else:
        grade = 'Low Load'

    return {
        'efficiency_score': total_score,
        'efficiency_grade': grade,
        'efficiency_breakdown': {
            'workload_pts': s_load,
            'ops_pts': s_ops,
            'speed_pts': s_speed,
            'attendance_pts': s_days,
            'has_discrepancy': diff > 0,
            'discrepancy_value': diff,
        },
    }


def _enrich(x: dict, detail: bool = False):
    """Enrich a cashier report row in-place and return (x, ops_dict)."""
    ops_cnt = float(x.get('operations_count', 0) or 0)
    ops_min = float(x.get('operations_minutes', 0) or 0)

    x['avg_seconds_per_operation'] = round(ops_min * 60 / ops_cnt, 1) if ops_cnt else 0
    x['operations_per_day'] = round(ops_cnt / x['days_worked'], 1) if x.get('days_worked') else 0

    raw = json.loads(x.get('metrics_json') or '{}')
    ops = raw.get('operations', {})
    x['employee_number'] = raw.get('employee_number', '')
    x['branch_name'] = clean_branch_name(x.get('branch_name', ''))

    # Load percent: prefer stored value from metrics_json, then from column
    lp = float(raw.get('load_percent') or x.get('load_percent') or 0)
    if 0 < lp <= 1.0:
        lp = round(lp * 100, 1)
    if lp == 0 and x.get('days_worked') and ops_min > 0:
        lp = round((ops_min / (x['days_worked'] * 480)) * 100, 1)
    x['load_percent'] = round(lp, 1)
    x['load_difference'] = float(raw.get('load_difference', 0) or 0)

    x['cashier_type'] = cashier_role(x)

    total_ops = ops_cnt or 1
    bek_cnt = float(x.get('bek_count', 0) or 0)
    front_cnt = float(x.get('front_count', 0) or 0)
    x['bek_pct'] = min(100.0, round((bek_cnt / total_ops) * 100, 1)) if ops_cnt else 0.0
    x['front_pct'] = min(100.0, round((front_cnt / total_ops) * 100, 1)) if ops_cnt else 0.0
    x['days_worked_pct'] = round((float(x.get('days_worked', 0) or 0) / 22) * 100, 1)

    eff = compute_efficiency_score(x)
    x['efficiency_score'] = eff['efficiency_score']
    x['efficiency_grade'] = eff['efficiency_grade']
    x['efficiency_breakdown'] = eff['efficiency_breakdown']

    x['hours_worked'] = round(ops_min / 60, 1)
    mins = int(ops_min)
    x['hours_str'] = f"{mins // 60} ч {mins % 60} мин"

    _EXCLUDE_KEYWORDS = ('жамибек', 'jamibek', 'жамифронт', 'jamfront',
                          'жамиамали', 'bekfark', 'frontfark', 'жами')
    real_metrics = {
        n: v for n, v in ops.items()
        if not any(k in norm(n).replace(' ', '') for k in _EXCLUDE_KEYWORDS)
    }
    real_ops_total = sum(v.get('count', 0) for v in real_metrics.values()) or ops_cnt or 1


    x['metrics'] = [
        {
            'name': n,
            'section': ('БЭК-операции' if _classify_op_group(n) == 'back' else 'ФРОНТ-операции'),
            'count': v.get('count', 0),
            'minutes': v.get('minutes', 0),
            'pct': round(v.get('count', 0) / real_ops_total * 100, 1),
        }
        for n, v in real_metrics.items()
    ]
    x['metrics'].sort(key=lambda m: m['count'], reverse=True)
    x['top_direction'] = x['metrics'][0]['name'] if x['metrics'] else '—'
    return x, ops


# ── Main analytics query ────────────────────────────────────────────────────

def cashier_analytics(
    import_id=None, page=1, page_size=25,
    role=None, search=None, position=None, status=None, branch=None,
):
    init_cashier_tables()
    with _connect() as c:
        if import_id is None:
            q = c.execute('SELECT id FROM cashier_imports ORDER BY id DESC LIMIT 1').fetchone()
            import_id = q['id'] if q else None

        info = None
        rows = []
        if import_id:
            info_row = c.execute('SELECT * FROM cashier_imports WHERE id=?', (import_id,)).fetchone()
            if info_row:
                info = dict(info_row)
                rows = [dict(z) for z in c.execute(
                    'SELECT * FROM cashier_reports WHERE import_id=? ORDER BY operations_count DESC',
                    (import_id,)
                )]

        # Fetch latest cashier statuses
        st_import = c.execute(
            'SELECT id, filename, imported_at FROM cashier_status_imports ORDER BY id DESC LIMIT 1'
        ).fetchone()
        status_map = {}
        status_map_2word = {}
        unmatched_statuses = []
        if st_import:
            if not info:
                info = {
                    'id': 0, 'filename': st_import['filename'],
                    'imported_at': st_import['imported_at'], 'rows_count': 0,
                }
            st_rows = c.execute(
                'SELECT * FROM cashier_statuses WHERE import_id=?', (st_import['id'],)
            ).fetchall()
            for s in st_rows:
                s_dict = dict(s)
                n_fn = norm(s_dict['full_name'])
                status_map[n_fn] = s_dict
                words = n_fn.split()
                if len(words) >= 2:
                    status_map_2word[f"{words[0]} {words[1]}"] = s_dict
                unmatched_statuses.append(s_dict)

        if not info:
            return {
                'summary': {}, 'cashiers': [], 'categories': [], 'positions': [],
                'top_by_position': {}, 'import': None, 'page': 1, 'total_pages': 1, 'total': 0,
            }

    matched_status_names = set()
    for x in rows:
        _enrich(x)
        n_fn = norm(x['full_name'])
        words = n_fn.split()
        st = status_map.get(n_fn)
        if not st and len(words) >= 2:
            st = status_map_2word.get(f"{words[0]} {words[1]}")
        if st:
            matched_status_names.add(norm(st['full_name']))
            x['hr_status_code'] = st['status_code']
            x['hr_status_label'] = st['status_label']
            x['branch_name'] = clean_branch_name(st.get('branch_name') or x.get('branch_name') or '')
            x['replacing_full_name'] = st.get('replacing_full_name')
            x['replaced_by_full_name'] = st.get('replaced_by_full_name')
            x['has_replacement'] = st.get('has_replacement', 0)
        else:
            x.setdefault('hr_status_code', 'active')
            x.setdefault('hr_status_label', '🟢 Работает')
            x['branch_name'] = clean_branch_name(x.get('branch_name') or '')

    # Add absent (maternity/vacation/sick) employees from shtat who have no KPI record.
    # Active employees NOT in KPI are ignored — they may be recent hires.
    # Only add employees with cashier-type positions (kassir/kassa).
    KASSIR_KEYWORDS = ('kassir', 'кассир', 'kassa', 'касса', 'gazna', 'g\'azna')
    for st in unmatched_statuses:
        if norm(st['full_name']) not in matched_status_names:
            st_code = st.get('status_code', 'active')
            if st_code not in ('maternity', 'vacation', 'sick'):
                continue
            st_pos = st.get('position') or ''
            n_pos = norm(st_pos)
            if not any(kw in n_pos for kw in KASSIR_KEYWORDS):
                continue
            dummy_row = {
                'id': 900000 + st['id'],
                'import_id': import_id,
                'tab_number': '—',
                'full_name': st['full_name'],
                'position': st_pos if st_pos else 'Кассир',
                'days_worked': 0,
                'operations_count': 0,
                'operations_minutes': 0,
                'bek_count': 0,
                'bek_minutes': 0,
                'front_count': 0,
                'front_minutes': 0,
                'metrics_json': '{}',
                'hr_status_code': st_code,
                'hr_status_label': st['status_label'],
                'branch_name': clean_branch_name(st.get('branch_name') or ''),
                'replacing_full_name': st.get('replacing_full_name'),
                'replaced_by_full_name': st.get('replaced_by_full_name'),
                'has_replacement': st.get('has_replacement', 0),
            }
            _enrich(dummy_row)
            rows.append(dummy_row)



    all_rows_list = list(rows)

    # Top 10 per position (pre-filter)
    pos_groups: dict[str, list] = {}
    for x in all_rows_list:
        pos = (x.get('position') or '').strip() or 'Прочее'
        pos_groups.setdefault(pos, []).append(x)

    top_by_position = {
        p: sorted(lst, key=lambda z: z.get('operations_count', 0), reverse=True)[:10]
        for p, lst in pos_groups.items()
    }
    positions_list = sorted(pos_groups.keys())

    # ── Filters ─────────────────────────────────────────────────────────────
    if role and str(role).strip():
        r_q = str(role).strip().lower()
        if r_q in ('front', 'back'):
            rows = [x for x in rows if x['cashier_type'] == r_q]

    if status and str(status).strip():
        st_q = str(status).strip().lower()
        if st_q == 'no_replacement':
            rows = [x for x in rows if x.get('has_replacement') == 0 and x.get('hr_status_code') in ('vacation', 'maternity')]
        elif st_q in ('active', 'vacation', 'maternity', 'temporary', 'vacant', 'sick'):
            rows = [x for x in rows if x.get('hr_status_code') == st_q]

    if position and str(position).strip():
        pos_q = str(position).strip().lower()
        rows = [x for x in rows if (x.get('position') or '').strip().lower() == pos_q]

    if search and str(search).strip():
        q_words = norm(search).split()
        if q_words:
            rows = [
                x for x in rows
                if all(
                    w in norm(
                        f"{x.get('full_name', '')} {x.get('position', '')} "
                        f"{x.get('tab_number', '')} {x.get('employee_number', '')} "
                        f"{x.get('hr_status_label', '')} {x.get('branch_name', '')} "
                        f"{x.get('raw_note', '')}"
                    )
                    for w in q_words
                )
            ]

    branches_list = sorted({
        (x.get('branch_name') or '').strip()
        for x in all_rows_list if (x.get('branch_name') or '').strip()
    })

    if branch and str(branch).strip():
        b_q = str(branch).strip().lower()
        rows = [x for x in rows if b_q in (x.get('branch_name') or '').strip().lower()]

    # ── Category aggregation ─────────────────────────────────────────────────
    cats: dict[str, dict] = {}
    allowed = BACK_GROUPS if role == 'back' else FRONT_GROUPS if role == 'front' else None
    bek_tot = 0.0
    front_tot = 0.0

    for x in rows:
        ops_data = json.loads(x.get('metrics_json') or '{}').get('operations', {})
        for n, v in ops_data.items():
            n_norm = norm(n).replace(' ', '')
            # Skip summary/excluded groups by keyword
            if any(k in n_norm for k in ('жамибек', 'jamibek', 'жамифронт', 'jamfront',
                                          'жамиамали', 'beкфарк', 'фронтфарк', 'bekfark',
                                          'frontfark', 'жами')):
                continue
            cnt = v.get('count', 0)
            mins = v.get('minutes', 0)
            # Classify by keyword
            op_group = _classify_op_group(n)
            if op_group == 'back':
                bek_tot += cnt
            elif op_group == 'front':
                front_tot += cnt
            if allowed is not None and op_group != (role or ''):
                continue
            a = cats.setdefault(n, {'name': n, 'count': 0, 'minutes': 0})
            a['count'] += cnt
            a['minutes'] += mins


    tot_cat_count = sum(c['count'] for c in cats.values()) or 1
    categories_list = []
    for c_item in sorted(cats.values(), key=lambda item: item['count'], reverse=True):
        c_item['pct'] = round(c_item['count'] / tot_cat_count * 100, 1)
        categories_list.append(c_item)

    total = len(rows)
    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    start = (page - 1) * page_size

    total_ops = bek_tot + front_tot
    total_min = sum(float(x.get('operations_minutes', 0) or 0) for x in rows)
    tot_load = sum(float(x.get('load_percent', 0) or 0) for x in rows)
    tot_eff = sum(float(x.get('efficiency_score', 0) or 0) for x in rows)

    summary = {
        'cashiers': total,
        'operations': round(total_ops),
        'minutes': round(total_min),
        'hours': round(total_min / 60, 1),
        'avg_seconds_per_operation': round(total_min * 60 / total_ops, 1) if total_ops else 0,
        'avg_load_percent': round(tot_load / total, 1) if total else 0,
        'avg_efficiency_score': round(tot_eff / total, 1) if total else 0,
        'bek_operations': round(bek_tot),
        'bek_pct': round(bek_tot / total_ops * 100, 1) if total_ops else 0,
        'front_operations': round(front_tot),
        'front_pct': round(front_tot / total_ops * 100, 1) if total_ops else 0,
        'back_cashiers': sum(1 for x in rows if x.get('cashier_type') == 'back'),
        'front_cashiers': sum(1 for x in rows if x.get('cashier_type') == 'front'),
        'active_cashiers': sum(1 for x in rows if x.get('hr_status_code') == 'active'),
        'vacation_cashiers': sum(1 for x in rows if x.get('hr_status_code') == 'vacation'),
        'maternity_cashiers': sum(1 for x in rows if x.get('hr_status_code') == 'maternity'),
        'temporary_cashiers': sum(1 for x in rows if x.get('hr_status_code') == 'temporary'),
        'sick_cashiers': sum(1 for x in rows if x.get('hr_status_code') == 'sick'),
        'vacant_positions': sum(1 for x in rows if x.get('hr_status_code') == 'vacant'),
        'no_replacement_cashiers': sum(
            1 for x in rows
            if x.get('has_replacement') == 0 and x.get('hr_status_code') in ('vacation', 'maternity')
        ),
        'active_filter': role or 'all',
    }

    return {
        'import': info,
        'summary': summary,
        'cashiers': rows[start:start + page_size],
        'categories': categories_list,
        'positions': positions_list,
        'branches': branches_list,
        'top_by_position': top_by_position,
        'page': page,
        'page_size': page_size,
        'total': total,
        'total_pages': max(1, (total + page_size - 1) // page_size),
    }


# ── Single cashier detail ────────────────────────────────────────────────────

def cashier_detail(report_id: int):
    init_cashier_tables()
    if report_id >= 900000:
        st_id = report_id - 900000
        with _connect() as c:
            st = c.execute('SELECT * FROM cashier_statuses WHERE id=?', (st_id,)).fetchone()
            if not st:
                return None
            st_info = dict(st)
            r = {
                'id': report_id,
                'import_id': 0,
                'tab_number': '—',
                'full_name': st_info['full_name'],
                'position': st_info.get('position') or 'Кассир',
                'days_worked': 0,
                'operations_count': 0,
                'operations_minutes': 0,
                'bek_count': 0,
                'bek_minutes': 0,
                'front_count': 0,
                'front_minutes': 0,
                'metrics_json': '{}',
                'filename': 'Реестр штата',
                'imported_at': datetime.now(timezone.utc).isoformat(),
            }
        x, _ = _enrich(dict(r), True)
        x['rank_by_operations'] = '—'
        x['percentile'] = 0
        x['hr_status_code'] = st_info['status_code']
        x['hr_status_label'] = st_info['status_label']
        x['branch_name'] = st_info.get('branch_name', '')
        x['replacing_full_name'] = st_info.get('replacing_full_name')
        x['replaced_by_full_name'] = st_info.get('replaced_by_full_name')
        x['has_replacement'] = st_info.get('has_replacement', 0)
        return x

    with _connect() as c:
        r = c.execute(
            'SELECT r.*, i.filename, i.imported_at FROM cashier_reports r '
            'JOIN cashier_imports i ON i.id=r.import_id WHERE r.id=?',
            (report_id,)
        ).fetchone()
        if not r:
            return None
        peers = c.execute(
            'SELECT operations_count FROM cashier_reports WHERE import_id=?',
            (r['import_id'],)
        ).fetchall()

        st_import = c.execute(
            'SELECT id FROM cashier_status_imports ORDER BY id DESC LIMIT 1'
        ).fetchone()
        st_info = None
        if st_import:
            st = c.execute(
                'SELECT * FROM cashier_statuses WHERE import_id=? AND (full_name=? OR lower(full_name)=lower(?))',
                (st_import['id'], r['full_name'], r['full_name'])
            ).fetchone()
            if st:
                st_info = dict(st)

    x, _ = _enrich(dict(r), True)
    tot_peers = len(peers) or 1
    x['rank_by_operations'] = 1 + sum(p['operations_count'] > x['operations_count'] for p in peers)
    x['percentile'] = round(100 * sum(p['operations_count'] <= x['operations_count'] for p in peers) / tot_peers, 1)

    if st_info:
        x['hr_status_code'] = st_info['status_code']
        x['hr_status_label'] = st_info['status_label']
        x['branch_name'] = clean_branch_name(st_info.get('branch_name', ''))
        x['replacing_full_name'] = st_info.get('replacing_full_name')
        x['replaced_by_full_name'] = st_info.get('replaced_by_full_name')
        x['has_replacement'] = st_info.get('has_replacement', 0)
    else:
        x['hr_status_code'] = x.get('hr_status_code') or 'active'
        x['hr_status_label'] = x.get('hr_status_label') or '🟢 Работает'
        x['branch_name'] = clean_branch_name(x.get('branch_name') or '')
        x['replacing_full_name'] = x.get('replacing_full_name')
        x['replaced_by_full_name'] = x.get('replaced_by_full_name')
        x['has_replacement'] = x.get('has_replacement', 1)

    return x
